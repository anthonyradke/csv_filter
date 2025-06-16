import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import os
import pandas as pd
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter







def round_to_15_min(dt):
    discard = timedelta(minutes=dt.minute % 15, seconds=dt.second, microseconds=dt.microsecond)
    dt -= discard
    if discard >= timedelta(minutes=7.5):
        dt += timedelta(minutes=15)
    return dt

def simplify_name(full_name):
    if isinstance(full_name, str):
        if "|dac-" in full_name:
            return full_name.split("|")[0].split(".")[-1]
        elif ".FLN_" in full_name:
            return full_name.split(".FLN_", 1)[1]
        else:
            return full_name.split(".")[-1]
    return full_name

def ensure_unique_columns(df):
    seen = {}
    new_cols = []
    for col in df.columns:
        if col not in seen:
            seen[col] = 1
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    df.columns = new_cols
    return df

def process_file_gui(file_path, log_box):
    filename = os.path.basename(file_path)
    log_box.insert(tk.END, f"\n📄 Processing: {filename}\n")
    log_box.update()

    try:
        df_raw = pd.read_csv(file_path, header=None, skiprows=[0, 2])
        clean_df = pd.DataFrame()
        skipped_cols = []

        for i in range(0, df_raw.shape[1], 4):
            try:
                time_col = pd.to_datetime(df_raw.iloc[1:, i].astype(str), errors='coerce')
                values = df_raw.iloc[1:, i + 1]
                time_rounded = time_col.map(round_to_15_min)
                valid = time_rounded.notna()
                title = str(df_raw.iloc[0, i])

                temp_df = pd.DataFrame({
                    "datetime": time_rounded[valid],
                    title: values[valid]
                })

                clean_df = temp_df if clean_df.empty else pd.merge(clean_df, temp_df, on="datetime", how="outer")
            except Exception:
                skipped_cols.append(get_column_letter(i + 1))

        if skipped_cols:
            joined = ", ".join(skipped_cols)
            log_box.insert(tk.END, f"⚠️  Skipped groups at columns: {joined}\n")
            log_box.update()

        if "datetime" not in clean_df.columns:
            raise ValueError("'datetime' column missing after processing.")

        clean_df = clean_df.groupby("datetime", as_index=False).first().sort_values("datetime").reset_index(drop=True)
        clean_df = clean_df.rename(columns={col: simplify_name(col) for col in clean_df.columns if col != "datetime"})
        return ensure_unique_columns(clean_df)
    except Exception as e:
        log_box.insert(tk.END, f"❌ Failed to process: {e}\n")
        log_box.update()
        return None

def start_gui():
    def browse_files():
        files = filedialog.askopenfilenames(filetypes=[("CSV Files", "*.csv")])
        if files:
            selected_files.clear()
            selected_files.extend(files)
            file_list_var.set("\n".join(os.path.basename(f) for f in files))

    def browse_folder():
        folder = filedialog.askdirectory()
        if folder:
            output_folder_var.set(folder)

    def reset_fields():
        selected_files.clear()
        file_list_var.set("")
        output_folder_var.set("")
        filename_entry.delete(0, tk.END)
        log_box.delete(1.0, tk.END)

    def update_filename_visibility(*args):
        if "Merge" in merge_var.get():
            filename_entry.config(state="normal")
        else:
            filename_entry.config(state="disabled")

    def start_processing():
        if not selected_files:
            messagebox.showwarning("No files selected", "Please select one or more CSV files.")
            return

        out_dir = output_folder_var.get() or os.path.dirname(selected_files[0])
        out_format = format_var.get()
        mode = merge_var.get()

        if "Merge" in mode:
            output_name = filename_entry.get().strip()
            if not output_name:
                messagebox.showerror("Missing Filename", "Please enter a name for the merged output file.")
                return
            save_path = os.path.join(out_dir, f"{output_name}.{out_format}")

        log_box.delete(1.0, tk.END)
        processed_data = {}
        for file_path in selected_files:
            df = process_file_gui(file_path, log_box)
            if df is not None:
                processed_data[os.path.splitext(os.path.basename(file_path))[0]] = df

        if not processed_data:
            return messagebox.showerror("Error", "No valid data to save.")

        if mode == "Keep files separate":
            for name, df in processed_data.items():
                save_path = os.path.join(out_dir, f"{name}_Filtered.{out_format}")
                if out_format == "csv":
                    df.to_csv(save_path, index=False)
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Cleaned Data"
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 31.00
                    for cell in ws['A'][1:]:
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                    wb.save(save_path)
                log_box.insert(tk.END, f"✅ Saved as: {save_path}\n")
        else:
            if mode == "Merge into one sheet":
                all_df = pd.concat(processed_data.values(), axis=0, ignore_index=True)
                all_df = all_df.groupby("datetime", as_index=False).first().sort_values("datetime").reset_index(drop=True)
                all_df = ensure_unique_columns(all_df)

                if out_format == "csv":
                    all_df.to_csv(save_path, index=False)
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Combined Data"
                    for r in dataframe_to_rows(all_df, index=False, header=True):
                        ws.append(r)
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 31.00
                    for cell in ws['A'][1:]:
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                    wb.save(save_path)
            else:
                wb = Workbook()
                wb.remove(wb.active)
                if "master" in mode.lower():
                    all_df = pd.concat(processed_data.values(), axis=0, ignore_index=True)
                    all_df = all_df.groupby("datetime", as_index=False).first().sort_values("datetime").reset_index(drop=True)
                    all_df = ensure_unique_columns(all_df)
                    ws = wb.create_sheet("Master Sheet")
                    for r in dataframe_to_rows(all_df, index=False, header=True):
                        ws.append(r)
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 31.00
                    for cell in ws['A'][1:]:
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                for name, df in processed_data.items():
                    ws = wb.create_sheet(name[:31])
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                    for col in ws.columns:
                        ws.column_dimensions[col[0].column_letter].width = 31.00
                    for cell in ws['A'][1:]:
                        cell.number_format = 'm/d/yyyy h:mm:ss AM/PM'
                wb.save(save_path)
            log_box.insert(tk.END, f"✅ Combined file saved as: {save_path}\n")

        messagebox.showinfo("Done", "Batch processing complete!")

    root = tk.Tk()
    root.title("Super Cool Andrew Tool")

    window_width = 860
    window_height = 620
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x_coord = (screen_width // 2) - (window_width // 2)
    y_coord = (screen_height // 2) - (window_height // 2)
    root.geometry(f"{window_width}x{window_height}+{x_coord}+{y_coord}")
    root.resizable(True, True)

    notebook = ttk.Notebook(root)
    notebook.pack(fill="both", expand=True)

    # === Main Tab ===
    main_tab = tk.Frame(notebook)
    notebook.add(main_tab, text="Main")

    selected_files = []
    file_list_var = tk.StringVar()
    output_folder_var = tk.StringVar()
    format_var = tk.StringVar(value="xlsx")
    merge_var = tk.StringVar(value="Keep files separate")

    tk.Label(main_tab, text="1. Select CSV File(s):").pack(anchor="w", padx=10, pady=(10, 0))
    tk.Button(main_tab, text="Browse CSV Files", command=browse_files).pack(anchor="w", padx=10)
    tk.Label(main_tab, textvariable=file_list_var, justify="left", fg="gray", wraplength=700).pack(anchor="w", padx=20)

    tk.Label(main_tab, text="2. Select Output Format:").pack(anchor="w", padx=10, pady=(10, 0))
    ttk.Combobox(main_tab, textvariable=format_var, values=["xlsx", "csv"], state="readonly", width=10).pack(anchor="w", padx=10)
    tk.Label(main_tab, text="Note: XLSX supports column formatting, CSV does not.", fg="gray").pack(anchor="w", padx=20)

    tk.Label(main_tab, text="3. Select Output Mode:").pack(anchor="w", padx=10, pady=(10, 0))
    ttk.Combobox(main_tab, textvariable=merge_var, values=[
        "Keep files separate",
        "Merge into one file with one sheet",
        "Merge into one file with separate sheets",
        "Merge into master file: one file, master sheet, and individual sheets"
    ], state="readonly", width=50).pack(anchor="w", padx=10)
    merge_var.trace_add("write", update_filename_visibility)

    tk.Label(main_tab, text="4. Output file name (if merging):").pack(anchor="w", padx=10, pady=(10, 0))
    filename_entry = tk.Entry(main_tab, width=50)
    filename_entry.pack(anchor="w", padx=10)
    filename_entry.config(state="disabled")

    tk.Label(main_tab, text="5. Optional Output Folder:").pack(anchor="w", padx=10, pady=(10, 0))
    tk.Button(main_tab, text="Browse Output Folder", command=browse_folder).pack(anchor="w", padx=10)
    tk.Label(main_tab, textvariable=output_folder_var, justify="left", fg="gray").pack(anchor="w", padx=20)

    start_button = tk.Button(
        main_tab,
        text="Start Processing",
        command=start_processing,
        bg="#4CAF50",       # green background
        fg="black",         # text color
        activebackground="#45a049",  # on click
        activeforeground="white",    # text on click
        relief="raised",    # gives it visible depth
        borderwidth=2
    )
    start_button.pack(anchor="w", padx=10, pady=(10, 5))
    reset_button = tk.Button(main_tab, text="Reset All Fields", command=reset_fields, bg="#f44336", fg="white")
    reset_button.place(relx=1.0, y=10, anchor="ne", x=-10)

    # === Log Tab ===
    log_tab = tk.Frame(notebook)
    notebook.add(log_tab, text="Console Log")
    log_box = scrolledtext.ScrolledText(log_tab, height=30, width=110)
    log_box.pack(padx=10, pady=10, fill="both", expand=True)
    log_box.bind_all("<MouseWheel>", lambda e: log_box.yview_scroll(-1*(e.delta//120), "units"))

    root.mainloop()

start_gui()
